
# import module
from openpyxl import Workbook, load_workbook
  
# load excel with its path
wrkbk = load_workbook("learners_database.xlsx")
  
wrksh = wrkbk.active
 
# iterate through excel and display data
for i in range(1, wrksh.max_row+1):
    print("\n")
    print("Row ", i, " data :")
      
    for j in range(1, wrksh.max_column+1):
        cell_obj = wrksh.cell(row=i, column=j)
        name = cell_obj.value
        # print(name)
        print(cell_obj.value, end=" ")